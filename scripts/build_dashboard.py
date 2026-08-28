#!/usr/bin/env python3
"""
NetSage AI - Dashboard Builder

Reads data/cases.csv and data/review_log.csv and writes outputs/dashboard.xlsx
with three tabs (Dashboard, Cases, Review) and three charts.

Usage:
    pip install openpyxl
    python scripts/build_dashboard.py
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT          = Path(__file__).parent.parent
CASES_CSV     = ROOT / "data" / "cases.csv"
REVIEW_CSV    = ROOT / "data" / "review_log.csv"
OUTPUT_XLSX   = ROOT / "outputs" / "dashboard.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

# ── Styles ────────────────────────────────────────────────────────────────────
FONT         = "Arial"
HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT  = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name=FONT, bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
LABEL_FONT   = Font(name=FONT, bold=True, size=11)
BODY_FONT    = Font(name=FONT, size=11)
THIN         = Side(style="thin", color="D9D9D9")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# ─────────────────────────────────────────────────────────────────────────────


def load_data() -> tuple[list[dict], list[dict]]:
    with open(CASES_CSV, newline="", encoding="utf-8") as f:
        cases = list(csv.DictReader(f))
    with open(REVIEW_CSV, newline="", encoding="utf-8") as f:
        review = list(csv.DictReader(f))
    return cases, review


def build_cases_sheet(wb: Workbook, cases: list[dict]) -> int:
    ws = wb.active
    ws.title = "Cases"
    fields = ["case_id", "category", "symptom", "expected_fault", "osi_layer", "severity"]
    ws.append(fields)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for row in cases:
        ws.append([row[k] for k in fields])
    for col, w in {"A": 10, "B": 14, "C": 55, "D": 60, "E": 12, "F": 10}.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
    ws.freeze_panes = "A2"
    return len(cases)


def build_review_sheet(wb: Workbook, review: list[dict]) -> int:
    ws = wb.create_sheet("Review")
    fields = ["case_id", "ai_root_cause", "ai_confidence", "human_verdict",
              "corrected_root_cause", "reviewer_notes"]
    ws.append(fields)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for row in review:
        ws.append([row[k] for k in fields])
    for col, w in {"A": 10, "B": 55, "C": 12, "D": 14, "E": 55, "F": 55}.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
    ws.freeze_panes = "A2"
    return len(review)


def build_dashboard_sheet(wb: Workbook, cases: list[dict], n_cases: int, n_rev: int):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "NetSage AI — Case Dashboard"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Auto-calculated from the Cases and Review sheets — edit those tabs and this recalculates."
    ws["B3"].font = SUBTITLE_FONT

    # Table 1: Issue type counts
    ws["B5"] = "Cases by Issue Type"
    ws["B5"].font = LABEL_FONT
    categories = sorted({r["category"] for r in cases})
    r0 = 6
    ws.cell(row=r0, column=2, value="Category").font = HEADER_FONT
    ws.cell(row=r0, column=2).fill = HEADER_FILL
    ws.cell(row=r0, column=3, value="Count").font = HEADER_FONT
    ws.cell(row=r0, column=3).fill = HEADER_FILL
    for i, cat in enumerate(categories, start=1):
        row = r0 + i
        ws.cell(row=row, column=2, value=cat).font = BODY_FONT
        ws.cell(row=row, column=3, value=f"=COUNTIF(Cases!$B$2:$B${n_cases+1},B{row})").font = BODY_FONT
    cat_last_row = r0 + len(categories)

    # Table 2: Severity counts
    sev_col = 6
    ws.cell(row=r0 - 1, column=sev_col, value="Cases by Severity").font = LABEL_FONT
    severities = ["Critical", "High", "Medium", "Low"]
    ws.cell(row=r0, column=sev_col, value="Severity").font = HEADER_FONT
    ws.cell(row=r0, column=sev_col).fill = HEADER_FILL
    ws.cell(row=r0, column=sev_col + 1, value="Count").font = HEADER_FONT
    ws.cell(row=r0, column=sev_col + 1).fill = HEADER_FILL
    for i, sev in enumerate(severities, start=1):
        row = r0 + i
        ws.cell(row=row, column=sev_col, value=sev).font = BODY_FONT
        ws.cell(row=row, column=sev_col + 1,
                value=f"=COUNTIF(Cases!$F$2:$F${n_cases+1},{get_column_letter(sev_col)}{row})").font = BODY_FONT
    sev_last_row = r0 + len(severities)

    # Table 3: AI vs Human verdict
    agree_row0 = max(cat_last_row, sev_last_row) + 3
    ws.cell(row=agree_row0 - 1, column=2, value="AI vs Human Review Outcome").font = LABEL_FONT
    verdicts = ["Accepted", "Edited", "Rejected"]
    ws.cell(row=agree_row0, column=2, value="Verdict").font = HEADER_FONT
    ws.cell(row=agree_row0, column=2).fill = HEADER_FILL
    ws.cell(row=agree_row0, column=3, value="Count").font = HEADER_FONT
    ws.cell(row=agree_row0, column=3).fill = HEADER_FILL
    for i, v in enumerate(verdicts, start=1):
        row = agree_row0 + i
        ws.cell(row=row, column=2, value=v).font = BODY_FONT
        ws.cell(row=row, column=3,
                value=f"=COUNTIF(Review!$D$2:$D${n_rev+1},B{row})").font = BODY_FONT
    verdict_last_row = agree_row0 + len(verdicts)

    # KPIs
    kpi_row = verdict_last_row + 2
    ws.cell(row=kpi_row, column=2, value="Total cases:").font = LABEL_FONT
    ws.cell(row=kpi_row, column=3, value=f"=COUNTA(Cases!$A$2:$A${n_cases+1})").font = BODY_FONT
    ws.cell(row=kpi_row + 1, column=2, value="AI agreement rate (Accepted / Total):").font = LABEL_FONT
    ws.cell(row=kpi_row + 1, column=3,
            value=f"=C{agree_row0+1}/SUM(C{agree_row0+1}:C{verdict_last_row})").font = BODY_FONT
    ws.cell(row=kpi_row + 1, column=3).number_format = "0.0%"
    ws.cell(row=kpi_row + 2, column=2, value="Cases needing correction (Edited + Rejected):").font = LABEL_FONT
    ws.cell(row=kpi_row + 2, column=3,
            value=f"=SUM(C{agree_row0+2}:C{verdict_last_row})").font = BODY_FONT

    for col, w in {"A": 2, "B": 34, "C": 10, "D": 4, "E": 4, "F": 12, "G": 10}.items():
        ws.column_dimensions[col].width = w

    # Charts
    bar = BarChart()
    bar.title = "Cases by Issue Type"
    bar.y_axis.title = "Cases"
    bar.x_axis.title = "Category"
    bar.add_data(Reference(ws, min_col=3, min_row=r0, max_row=cat_last_row), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=r0 + 1, max_row=cat_last_row))
    bar.width, bar.height = 16, 8
    ws.add_chart(bar, "I5")

    pie = PieChart()
    pie.title = "Cases by Severity"
    pie.add_data(Reference(ws, min_col=sev_col + 1, min_row=r0, max_row=sev_last_row), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=sev_col, min_row=r0 + 1, max_row=sev_last_row))
    pie.width, pie.height = 12, 8
    ws.add_chart(pie, "I22")

    bar2 = BarChart()
    bar2.title = "AI vs Human Review Outcome"
    bar2.y_axis.title = "Cases"
    bar2.add_data(Reference(ws, min_col=3, min_row=agree_row0, max_row=verdict_last_row), titles_from_data=True)
    bar2.set_categories(Reference(ws, min_col=2, min_row=agree_row0 + 1, max_row=verdict_last_row))
    bar2.width, bar2.height = 12, 8
    ws.add_chart(bar2, "Q5")


def main():
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    cases, review = load_data()
    wb = Workbook()
    n_cases = build_cases_sheet(wb, cases)
    n_rev   = build_review_sheet(wb, review)
    build_dashboard_sheet(wb, cases, n_cases, n_rev)
    wb.save(OUTPUT_XLSX)
    print(f"Saved {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
